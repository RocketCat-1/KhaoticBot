import discord
from openpyxl import Workbook, load_workbook
from discord.ext import tasks
from datetime import datetime
import os
intents = discord.Intents.all()
bot = discord.Bot(intents=intents)

@bot.event
async def on_ready():
    if os.path.exists("settings.xlsx") == False:
        workbook = Workbook()
        workbook.save(filename="settings.xlsx")
    await bot.change_presence(activity=discord.Game(name="Khaotic Soulz's bot."))
    print(f"Logged in as {bot.user}")

@bot.slash_command(description="Blacklists mentioned user from using the bot.")
async def blacklist(ctx, user: discord.Member):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+2):
            admins.append(int(sheet.cell(row=x, column=2).value))
    if ctx.author.id in admins:
        if user.id in admins:
            await ctx.respond(str(user) + " is an admin and can't be blacklisted.")
            return
        number = str(int(sheet.cell(row=1, column=1).value)+1)
        sheet["B"+ number] = str(user.id)
        sheet["B1"] = number
        workbook.save(filename="settings.xlsx")
        await ctx.respond(str(user) + " is now blacklisted")
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Makes the mentioned user an admin.")
async def giveadmin(ctx, user: discord.Member):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
    if ctx.author.id in admins:
        if user in admins:
            await ctx.respond(str(user) + " is already an admin.")
            return
        number = str(int(sheet.cell(row=1, column=2).value)+1)
        sheet["B"+ number] = str(user.id)
        sheet["B1"] = number
        workbook.save(filename="settings.xlsx")
        await ctx.respond(str(user) + " is now an admin.")
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Changes bot's status.")
async def status(ctx, status):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
    if ctx.author.id in admins:
        await bot.change_presence(activity=discord.Game(name=status))
        await ctx.respond("Bot status changed.", ephemeral = True)
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)


@bot.slash_command(description="Removes the mentioned user from the admin list.")
async def removeadmin(ctx, user: discord.Member):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
            if int(sheet.cell(row=x, column=2).value) == user.id:
                userRow = x
    if ctx.author.id in admins:
        if user.id in admins: 
            number = str(int(sheet.cell(row=1, column=2).value))
            sheet["B"+ str(userRow)] = str(sheet.cell(row=int(number), column=2).value)
            sheet["B"+ number] = ""
            number = str(int(sheet.cell(row=1, column=2).value)-1)
            sheet["B1"] = number
            workbook.save(filename="settings.xlsx")
            await ctx.respond(str(user) + " is no longer an admin.")
        else:
            await ctx.respond(str(user) + " isn't an admin.")
            return
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Removes the mentioned user from the blacklist.")
async def unblacklist(ctx, user: discord.Member):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
            if int(sheet.cell(row=x, column=2).value) == user.id:
                userRow = x
    if ctx.author.id in admins:
        if user.id in admins: 
            number = str(int(sheet.cell(row=1, column=2).value))
            sheet["B"+ str(userRow)] = str(sheet.cell(row=int(number), column=2).value)
            sheet["B"+ number] = ""
            number = str(int(sheet.cell(row=1, column=2).value)-1)
            sheet["B1"] = number
            workbook.save(filename="settings.xlsx")
            await ctx.respond(str(user) + " is no longer an admin.")
        else:
            await ctx.respond(str(user) + " isn't an admin.")
            return
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Summons bot to current voice channel.")
async def connect(ctx):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
        workbook.close()
    if ctx.author.id in admins:
        voice_state = ctx.user.voice
        if voice_state is None:
            await ctx.response.send_message('You need to be in a voice channel to use this command')
            return
        await ctx.respond("Now announcing respawns in: " + str(ctx.user.voice.channel))
        vc = await ctx.user.voice.channel.connect()
        respawnLoop.start(vc, ctx)
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Disconnects the bot from the current voice channel.")
async def disconnect(ctx):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
        workbook.close()
    if ctx.author.id in admins:
        voice_state = ctx.user.voice
        if voice_state is None:
            await ctx.response.send_message('You need to be in a voice channel to use this command.')
            return
        if ctx.guild.voice_client in bot.voice_clients:
            memids = []
            channel = ctx.user.voice.channel
            for member in ctx.user.voice.channel.members:
                memids.append(str(member.id))
            if "1080825215833096202" in memids:
                await ctx.guild.voice_client.disconnect()
                await ctx.respond("Disconnect bot from channel: " + str(channel))
            else:
                await ctx.respond("The bot isn't in your current voice channel.", ephemeral = True)
        else:
            await ctx.respond("The bot isn't in your current server.", ephemeral = True)
    else:
        await ctx.respond("Only admins can perform this command.", ephemeral = True)

@bot.slash_command(description="Gives a list of the available commands to you.")
async def help(ctx):
    workbook = load_workbook(filename="settings.xlsx")
    sheet = workbook.active
    admins = []
    if int(sheet.cell(row=1, column=2).value) != 1:
        for x in range(2, int(sheet.cell(row=1, column=2).value)+1):
            admins.append(int(sheet.cell(row=x, column=2).value))
        workbook.close()
    if ctx.author.id in admins:
        embedVar = discord.Embed(title="Commands **[ADMIN]**", description="**/disconnect** - Disconnects the bot from your voice channel.\n**/connect** - Connects the bot to your current voice channel.\n**/unblacklist** ``user`` - Removes the mentioned user from the blacklist.\n**/blacklist** ``user`` - Blacklists a user from using the bot\n**/giveadmin** ``user`` - Adds the mentioned user to the admin list.\n**/removeadmin** ``user`` - Removes the mentioned user from the admin", color=0x00ff00)
        await ctx.respond(embed = embedVar, ephemeral = True)
    else:
        await ctx.respond("Currently all commands are admin commands, please check back in future updates or DM <@259726664236269568> for more info.", ephemeral = True)

@bot.slash_command(description="Gives a list of who has crafting skills at 200.")
async def crafters(ctx):
    weapon = []
    armor = []
    enginer = []
    jewel = []
    arcana = []
    cooking = []
    furnishing = []
    masters = []
    for member in ctx.guild.members:
        for role in member.roles:
            roles = 0
            if role.id == 1074886454620209202: #Weaponsmithing
                weapon.append(member.name)
                roles = roles + 1
            if role.id == 1074886497439846491: #Armoring
                armor.append(member.name)
                roles = roles + 1
            if role.id == 1074886525235507240: #Enginering 
                enginer.append(member.name)
                roles = roles + 1
            if role.id == 1074886552720777226: #Jewel
                jewel.append(member.name)
                roles = roles + 1
            if role.id == 1074886597914398780: #Arcana
                arcana.append(member.name)
                roles = roles + 1
            if role.id == 1074886623029891124: #Cooking
                cooking.append(member.name)
                roles = roles + 1
            if role.id == 1074886647054872677: #Furnishing 
                furnishing.append(member.name)
                roles = roles + 1
            if roles == 7:
                masters.append(member.name)
    embedVar = discord.Embed(title="Master Crafter List", description="**Masters:**\n" + ', '.join(masters) + "\n\n**Arcana**:" + ', '.join(arcana) + "\n\n**Armoring**:" + ', '.join(armor) + "\n\n**Cooking**:" + ', '.join(cooking) + "\n\n**Engineering**:" + ', '.join(enginer) + "\n\n**Furnishing**:" + ', '.join(furnishing) + "\n\n**Jewelcrafting**:" + ', '.join(jewel) + "\n\n**Weaponsmithing**:" + ', '.join(weapon), color=0x00ff00)
    await ctx.respond(embed = embedVar, ephemeral = True)

@tasks.loop(seconds=0.5)
async def respawnLoop(vc, ctx):
    if ctx.guild.voice_client not in bot.voice_clients:
        respawnLoop.stop()
        return
    currentDateAndTime = datetime.now()
    if currentDateAndTime.minute > 30:
        warminute = 59 - currentDateAndTime.minute
    else:   
        warminute = 29 - currentDateAndTime.minute
    if warminute <= 9:
        warminute = "0" + str(warminute)
    warsecond = 59 - currentDateAndTime.second
    if warsecond <= 9:
        warsecond = "0" + str(warsecond)
    times = ['2940', '2920', '2900', '2840', '2820', '2800', '2740', '2720', '2700', '2640', '2620', '2600', '2540', '2520', '2452', '2424', '2356', '2328', '2300', '2232', '2204', '2136', '2108', '2040 ', '2012 ', '1944 ', '1908', '1832', '1756', '1720', '1644', '1608', '1532', '1456', '1420', '1336', '1252', '1208', '1124', '1040', '0956', '0912', '0820 ', '0728 ', '0636', '0544', '0452', '0352', '0252', '0152', '0052']
    time = str(warminute) + str(warsecond)
    time = int(time)
    found = 0
    for x in range(0, len(times)):
        if time < int(times[x]) or found == 1:
            pass
        elif time >= int(times[x]):
            found = 1
            timestime = times[x]
            timeminute = timestime[0] + timestime[1]
            timeminute = int(timeminute)
            timesecond = timestime[2] + timestime[3]
            timesecond = int(timesecond)
            if int(warminute) - timeminute == 0 or timesecond >= 41:
                timersecond = int(warsecond) - timesecond
                if timersecond < 0:
                    timersecond = timersecond + 60
                if timersecond == 30 and is_playing(ctx) == False:
                    vc.play(discord.FFmpegPCMAudio(executable="C:/Program Files/ShareX/ffmpeg.exe", source="respawntimer.mp3"))
                elif timersecond == 20 and is_playing(ctx) == False:
                    vc.play(discord.FFmpegPCMAudio(executable="C:/Program Files/ShareX/ffmpeg.exe", source="respawntimer20seconds.mp3"))
                elif timersecond == 10 and is_playing(ctx) == False:
                    vc.play(discord.FFmpegPCMAudio(executable="C:/Program Files/ShareX/ffmpeg.exe", source="respawntimer10seconds.mp3"))

def is_playing(ctx):
    if ctx.voice_client.is_playing():
        return True
    else:
        return False
    
bot.run("bot token")
